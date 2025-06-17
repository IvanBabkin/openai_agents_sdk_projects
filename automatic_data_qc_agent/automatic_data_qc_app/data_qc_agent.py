from dotenv import load_dotenv
from agents import Agent, Runner, ModelSettings, function_tool
import pandas as pd
import PyPDF2
import io
import openpyxl
import json

load_dotenv(override=True)

INSTRUCTIONS = """
You are an expert data analyst reviewing an Excel file against PDF specifications. Provide a comprehensive analysis with precise issue locations and actionable recommendations.

## Analysis Scope

1. **Structure Compliance**: Verify sheets, columns, headers, data types, and field requirements match PDF specs
2. **Format Validation**: Check dates, numbers, text, codes, and boolean fields for correct formatting
3. **Data Consistency**: Validate temporal logic, numerical relationships, business rules, and data integrity
4. **Quality Checks**: Identify missing data, duplicates, outliers, and formatting issues

## For Each Issue Report

- **Location**: Sheet name, cell reference (e.g., "C15"), column name, row number
- **Type**: Format violation | Consistency error | Missing data | Invalid value | Structural problem
- **Severity**: Critical (blocks processing) | Major (significant deviation) | Minor (formatting)
- **Details**: Current value vs. expected value with clear explanation
- **Fix**: Specific Excel-based solution (no external code)

## Output Structure

EXECUTIVE SUMMARY
- Total issues by severity
- Key problem areas

DETAILED FINDINGS
Sheet: [Name]
Issue #1: [Type] - [Severity]
• Location: Cell B7, Column "Date"
• Found: "2024/13/45"
• Expected: Valid date (DD/MM/YYYY)
• Description: Invalid date format
• Fix: Correct date entry


## Key Instructions

- Compare against PDF examples/templates
- Prioritize data integrity issues
- Note where file exceeds requirements
- Flag unclear PDF sections
- Skip 'data dictionary' sheet
- Focus on Excel-native solutions only
- Make sure that any recommendations you make, can be done in Excel. Do not suggest the use of Python or any other code language to fix the identified issues.
- Ignore 'data dictionary' sheet. No analysis should be done on this sheet!
"""

# Global variables to store file content
_pdf_content = None
_excel_content = None

@function_tool
def read_pdf_specs() -> str:
    """Read PDF specification content that was previously loaded."""
    global _pdf_content
    if _pdf_content is None:
        return "PDF content not available. Please ensure the PDF file was properly loaded."
    return _pdf_content

@function_tool  
def read_excel_data_raw() -> str:
    """Read Excel data content that was previously loaded."""
    global _excel_content
    if _excel_content is None:
        return "Excel content not available. Please ensure the Excel file was properly loaded."
    return _excel_content

def load_pdf_content(pdf_bytes: bytes) -> str:
    """Load PDF content from bytes and store globally."""
    global _pdf_content
    try:
        pdf_stream = io.BytesIO(pdf_bytes)
        pdf_reader = PyPDF2.PdfReader(pdf_stream)
        content = []
        for page in pdf_reader.pages:
            content.append(page.extract_text())
        _pdf_content = "\n".join(content)
        return _pdf_content
    except Exception as e:
        _pdf_content = f"Error reading PDF: {str(e)}"
        return _pdf_content

def load_excel_content(excel_bytes: bytes) -> str:
    """Load Excel content from bytes and store globally."""
    global _excel_content
    try:
        excel_stream = io.BytesIO(excel_bytes)
        workbook = openpyxl.load_workbook(excel_stream, data_only=False)
        
        # Process all sheets except 'data dictionary' (as per instructions)
        all_sheets_data = {}
        
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == 'data dictionary':
                continue  # Skip data dictionary sheet as per instructions
                
            worksheet = workbook[sheet_name]
            
            structured_data = {
                "sheet_name": worksheet.title,
                "cells": {},
                "dimensions": {
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column
                }
            }
            
            for row in worksheet.iter_rows():
                for cell in row:
                    # Fix 1: Check for cell content properly
                    has_formula = cell.data_type == 'f'
                    has_value = cell.value is not None
                    
                    if has_value or has_formula:
                        cell_info = {
                            "coordinate": cell.coordinate,
                            "value": cell.value,
                            "data_type": cell.data_type,
                            "row": cell.row,
                            "column": cell.column
                        }
                        
                        # Fix 2: Check for formula properly
                        if has_formula:
                            cell_info["formula"] = cell.value  # The formula IS the value when data_type is 'f'
                        
                        # Include formatting if needed
                        if cell.number_format != 'General':
                            cell_info["number_format"] = cell.number_format
                        
                        structured_data["cells"][cell.coordinate] = cell_info
            
            all_sheets_data[sheet_name] = structured_data
        
        # FIX: Assign the processed data to the global variable
        _excel_content = json.dumps(all_sheets_data, indent=2, default=str)
        return _excel_content
    except Exception as e:
        _excel_content = f"Error reading Excel: {str(e)}"
        return _excel_content

def create_data_qc_agent():
    """Create and return the data QC agent."""
    return Agent(
        name="data_qc_agent",
        instructions=INSTRUCTIONS,
        tools=[read_pdf_specs, read_excel_data_raw], 
        model="gpt-4o-mini",
        model_settings=ModelSettings(
            tool_choice="required"
        )
    )

async def analyze_data_quality(pdf_bytes: bytes, excel_bytes: bytes) -> str:
    """
    Analyze data quality by comparing Excel file against PDF specifications.
    
    Args:
        pdf_bytes: PDF file content as bytes
        excel_bytes: Excel file content as bytes
    
    Returns:
        str: Analysis report in markdown format
    """
    # Load file contents into global variables
    load_pdf_content(pdf_bytes)
    load_excel_content(excel_bytes)
    
    agent = create_data_qc_agent()
    
    prompt = """
    Execute the following data quality analysis workflow:
    
    1. FIRST: Call read_pdf_specs to get the PDF specifications
    2. SECOND: Call read_excel_data_raw to analyze the Excel file
    3. THIRD: Compare findings and provide detailed quality assessment
    
    Start by using the read_pdf_specs tool now.
    """
    
    result = await Runner.run(
        agent, 
        prompt,
        max_turns=5
    )
    
    return result.final_output